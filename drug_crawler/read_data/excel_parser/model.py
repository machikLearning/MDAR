#-*-coding:utf-8-*-

from drug_crawler.read_data import model
import openpyxl
import os

#Excel파일 생성 클래스

class ExcelWriter:

    def __init__(self):
        self._file_path = "atc_combine_by_component.xlsx" #저장파일 경로지정
        if(not os.path.exists(self._file_path)): #파일이 없다면 생성해준다
            f = open(self._file_path,'w')
            f.close()

    def write_csv(self, number,fail_list):     #fail_list는 엑셀에 저장할 데이터 목록
        try:
            wb = openpyxl.load_workbook(self._file_path)
            sheet = wb.create_sheet(str(number))
        except:
            wb = openpyxl.workbook.Workbook()
            sheet = wb.active

        data_list = ["item_name","entp_name","cnsgn_manuf","etc_otc_ocde","class_no","chart","bar_code","material_name","ee_doc_id","ud_doc_id","nb_doc_id",
               "insert_file","storage_method","valid_term","reexam_target","reexam_date","pack_unit","edi_code","doc_text","permit_kind_name","entp_no","make_material_flag","newdrug_class_name",
               "induty_type","cancel_date","cancel_name","change_date","narcotic_kind_code","gbn_name","ee_doc_data","ud_doc_data","nb_doc_data","pn_doc_data","main_item_ingr","ingr_name","item_permit_date"]
        sheet.append(data_list)
        for fail in fail_list: # 데이터를 엑셀파일에 기록한다
            sheet.append(fail)
        wb.save(self._file_path)
        wb.close()




class ProductionCodeParser(model.AbsReadData):
    def __init__(self, file_path):
        self._file_path = file_path

    def read_data(self):
        result = []
        try:
            with open(self._file_path, "r", encoding="utf-8") as f:
                result = f.readlines()
                f.close()
        except Exception as ex:
            print(ex)
        return result


class DrugTemplateExcelParser(model.AbsReadData):
    def __init__(self, file_path):
        self._file_path = file_path


    def read_data(self):
        result = []
        try:
            excel_document = openpyxl.load_workbook(self._file_path)
            sheet = excel_document.get_sheet_by_name("20년7월1일_(25,608)")
            rows = list(sheet.rows)[18018:]
            for row in rows:
                row = row[:4]
                row_values = []
                for cell in row:
                    row_values.append(cell.value)
                result.append(row_values)
        except Exception as ex:
            print(ex)
        finally:
            excel_document.close()
        return result

    def read_main_code(self):
        result = set()
        try:
            excel_document = openpyxl.load_workbook(self._file_path)
            sheet = excel_document.get_sheet_by_name("20년7월1일_(25,608)")
            for row in sheet.rows:
                result.add(row[0].value)
        except Exception as ex:
            print(ex)
        finally:
            excel_document.close()
        result.remove('주성분코드')
        return result

    def read_basic_drug(self):
        result =[]
        try:
            excel_document = openpyxl.load_workbook(self._file_path)
            sheet = excel_document.get_sheet_by_name("Sheet1")
            for row in sheet.rows:
                row_value = []
                row = row[:10]
                for cell in row:
                    if cell.value:
                        value = cell.value
                        if isinstance(value,str):
                            value = value.replace("'","")
                        row_value.append(value)
                    else:
                        row_value.append('NULL')
                result.append(row_value)
        except Exception as ex:
            print(ex)
        finally:
            excel_document.close()
        del result[0]
        return result